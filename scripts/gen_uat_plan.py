#!/usr/bin/env python3
"""Generate docs/uat-plan.xlsx — UAT test cases for the Knowledge Fabric.

One row per step. Five sheets: A_Operator, B_Reviewer, C_Admin,
D_Auditor, E_Demo. Re-runnable: regenerates the file in-place so we
can update the plan without manual spreadsheet editing.

Columns:
    TC-ID · Test · Instructions · Expected outcome · Pass/Fail · Notes
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = REPO_ROOT / "docs" / "uat-plan.xlsx"


# -----------------------------------------------------------------------------
# Test case model
# -----------------------------------------------------------------------------
@dataclass
class Step:
    instructions: str
    expected: str


@dataclass
class TestCase:
    tc_id: str
    title: str
    pre: str  # preconditions — rendered as the Setup row
    steps: list[Step]
    pass_criteria: Optional[str] = None  # rolled up into a final row if present


# -----------------------------------------------------------------------------
# A — Operator workflows
# -----------------------------------------------------------------------------
A_OPERATOR: list[TestCase] = [
    TestCase(
        tc_id="TC-OP-01",
        title="Submit a HITL case end-to-end",
        pre=("Logged in as operator (e.g. anuj). SC-PP-008 chip is "
             "available in the operator console."),
        steps=[
            Step(
                instructions=("Type \"onboard new supplier Hemlock Precision "
                              "Castings\" in the prompt box and submit."),
                expected=("Agent picks SC-PP-008 with high confidence. "
                          "Clarifying question appears."),
            ),
            Step(
                instructions="Click Confirm on the clarifying question.",
                expected=("Case moves to Binding, then Awaiting Review. "
                          "Envelope shows ~13 facts across 3 stages."),
            ),
            Step(
                instructions=("Expand each stage block (Policy & scope, "
                              "Agent gathered, For your decision)."),
                expected=("Fact cards render with ontology_type + id + "
                          "source pills."),
            ),
            Step(
                instructions="Switch to the Reviewer role pill.",
                expected=("Approve / Reject / Request more info buttons "
                          "unlock."),
            ),
        ],
        pass_criteria=("Case reaches Awaiting Review with non-empty envelope; "
                       "no console errors; lineage panel shows append-only events."),
    ),
    TestCase(
        tc_id="TC-OP-02",
        title="Replay a closed case",
        pre="At least one completed case exists from a prior session.",
        steps=[
            Step(
                instructions=("Open a complete case from the case list. "
                              "Click Replay -> Approve."),
                expected=("A new case is spawned, bypasses HITL, "
                          "auto-decides."),
            ),
            Step(
                instructions="Open the newly spawned case.",
                expected=("decision_kind is \"approve\". Lineage references "
                          "the original case as a sibling."),
            ),
            Step(
                instructions="Confirm the relationship in the case list.",
                expected=("New case is independent (distinct case_id). "
                          "Both cases visible in the list."),
            ),
        ],
    ),
    TestCase(
        tc_id="TC-OP-03",
        title="Case spec — historical pin",
        pre=("An Admin has bumped SC-PP-008 from v1 to v2 since the case "
             "in TC-OP-01 was created."),
        steps=[
            Step(
                instructions=("Open the case from TC-OP-01. Click "
                              "StatusBar -> Case spec."),
                expected="The Case spec modal opens.",
            ),
            Step(
                instructions="Read the header pill next to the scenario id.",
                expected=("Pill reads \"Ran against v1\" (or whichever "
                          "version was current at case-creation time)."),
            ),
            Step(
                instructions="Click through the Scenario and Ontology tabs.",
                expected=("Content reflects v1 spec, not the live v2."),
            ),
        ],
        pass_criteria=("Pin survives even after scenario is edited. "
                       "Reviewer / auditor can defend the decision."),
    ),
]


# -----------------------------------------------------------------------------
# B — Reviewer workflows
# -----------------------------------------------------------------------------
B_REVIEWER: list[TestCase] = [
    TestCase(
        tc_id="TC-RV-01",
        title="HITL approve with load-bearing facts",
        pre="Case from TC-OP-01 is awaiting review.",
        steps=[
            Step(
                instructions="Click Approve on the active case.",
                expected=("Approve modal opens with the execute_message "
                          "preview."),
            ),
            Step(
                instructions=("Locate the \"Which facts were load-bearing? "
                              "optional pick up to 3\" picker section."),
                expected="One chip per fact across all stages renders.",
            ),
            Step(
                instructions=("Click 2 chips (e.g. SanctionsProximity, "
                              "CarrierExposure)."),
                expected=("Counter reads (2/3). Picked chips highlight "
                          "purple / show selected state."),
            ),
            Step(
                instructions="Click a third chip.",
                expected=("Counter reads (3/3). Remaining chips disable "
                          "with a cap-tooltip."),
            ),
            Step(
                instructions="Click Confirm execute.",
                expected=("Modal closes. Case moves to Complete / Executed."),
            ),
            Step(
                instructions=("Re-open the case via Case spec."),
                expected=("\"Reviewer marked these facts as load-bearing (3)\" "
                          "section renders the 3 picks in order."),
            ),
        ],
        pass_criteria=("Cap of 3 enforced; refs persisted; visible in Case spec."),
    ),
    TestCase(
        tc_id="TC-RV-02",
        title="Reject with rationale + load-bearing",
        pre="Case is in review_ready state.",
        steps=[
            Step(
                instructions="Click Reject.",
                expected="Rationale modal opens.",
            ),
            Step(
                instructions=("Pick a quick-pick reason "
                              "(e.g. \"Sanctions proximity too close\")."),
                expected="Textarea fills with the picked reason text.",
            ),
            Step(
                instructions="Pick 2 load-bearing facts.",
                expected="Counter reads (2/3).",
            ),
            Step(
                instructions="Click Continue.",
                expected=("Confirm screen displays. \"Load-bearing facts\" "
                          "section lists the 2 picks."),
            ),
            Step(
                instructions="Click Yes, reject.",
                expected=("Case completes with decision_kind=reject."),
            ),
            Step(
                instructions=("Curl GET /api/cases/{id}/highlights with an "
                              "admin token."),
                expected="Response returns the 2 picked refs in order.",
            ),
        ],
    ),
    TestCase(
        tc_id="TC-RV-03",
        title="Promoted-pattern advisory chip (Phase 3b)",
        pre=("Admin has promoted \"SC-PP-008 / reject / SanctionsProximity\" "
             "via TC-AD-04. A new SC-PP-008 case is started by an operator."),
        steps=[
            Step(
                instructions="Open the new SC-PP-008 case after binding completes.",
                expected=("\"Compounding loop * 1 promoted pattern matched\" "
                          "advisory banner appears above the evidence map."),
            ),
            Step(
                instructions="Read the banner contents.",
                expected=("REJECT pill + SanctionsProximity + \"Reviewers "
                          "cited SanctionsProximity in X of Y reject "
                          "decisions...\" + source-date footer."),
            ),
            Step(
                instructions=("Confirm the banner is advisory, not blocking, "
                              "by clicking Approve."),
                expected=("Approve flow proceeds unchanged. The reviewer is "
                          "not coerced by the chip."),
            ),
        ],
        pass_criteria=("Banner present; reviewer not coerced; chip metadata "
                       "matches the underlying source case count + share."),
    ),
    TestCase(
        tc_id="TC-RV-04",
        title="Request more info with follow-up reminder",
        pre="Case in review_ready.",
        steps=[
            Step(
                instructions="Click Request more info.",
                expected=("Same rationale modal as Reject opens. Header tone "
                          "is info-style not danger-style."),
            ),
            Step(
                instructions=("Pick a follow-up reminder (e.g. \"3 days\")."),
                expected="Follow-up chip highlights as selected.",
            ),
            Step(
                instructions=("Fill rationale, Continue, then Yes, send back."),
                expected=("Case completes with decision_kind="
                          "request_more_info. Closing message references the "
                          "follow-up window."),
            ),
        ],
    ),
]


# -----------------------------------------------------------------------------
# C — Admin workflows
# -----------------------------------------------------------------------------
C_ADMIN: list[TestCase] = [
    TestCase(
        tc_id="TC-AD-01",
        title="Edit scenario — Phase 1 versioning",
        pre="Logged in as admin.",
        steps=[
            Step(
                instructions=("Knowledge -> Scenarios tab -> click the pencil "
                              "on SC-PP-008."),
                expected=("Edit modal opens. Footer reads \"Currently v1 * "
                          "Saving creates v2\"."),
            ),
            Step(
                instructions="Change the title. Click Save changes.",
                expected=("Alert: \"Saved as v2. The prior content is "
                          "preserved as v1.\""),
            ),
            Step(
                instructions="Re-open the SC-PP-008 edit modal.",
                expected="Footer reads \"Currently v2 * Saving creates v3\".",
            ),
            Step(
                instructions="Click \"Version history ->\".",
                expected=("Versions modal lists v1 and v2 newest-first, each "
                          "with title + timestamp."),
            ),
            Step(
                instructions="Click the v1 row in the list.",
                expected=("Right pane shows v1's original title -- "
                          "NOT the new one."),
            ),
            Step(
                instructions="Click the v2 row quickly after v1.",
                expected=("Right pane clears v1's body first, then loads v2. "
                          "No stale flash (Phase 1 audit fix)."),
            ),
        ],
    ),
    TestCase(
        tc_id="TC-AD-02",
        title="Concurrent edits don't lose data",
        pre="Logged in as admin in two browser tabs (use an incognito window for tab B).",
        steps=[
            Step(
                instructions=("Tab A: open SC-PP-008 in the edit modal. "
                              "Note current version."),
                expected="Footer shows \"Currently vN\".",
            ),
            Step(
                instructions=("Tab B (incognito, same admin): open the same "
                              "scenario in its edit modal."),
                expected="Footer shows the same \"Currently vN\".",
            ),
            Step(
                instructions="Tab A: change title to \"X-A\". Click Save.",
                expected="Alert says vN+1.",
            ),
            Step(
                instructions=("Tab B (do NOT refresh): change title to "
                              "\"X-B\". Click Save."),
                expected=("Alert says vN+2 (server's authoritative version, "
                          "not vN+1)."),
            ),
            Step(
                instructions="Open Version history on either tab.",
                expected=("All four versions present (original + N+1 (X-A) + "
                          "N+2 (X-B)). Both edits preserved."),
            ),
        ],
        pass_criteria=("Phase 1 audit fix holds -- concurrent saves "
                       "serialize, no data loss."),
    ),
    TestCase(
        tc_id="TC-AD-03",
        title="Insights modal (Phase 3a)",
        pre=("Logged in as admin. At least the 26 backfilled highlight rows "
             "are in the DB (prod default)."),
        steps=[
            Step(
                instructions="Click StatusBar -> Insights.",
                expected="Modal opens with the title \"Recurring override drivers\".",
            ),
            Step(
                instructions="Scan the modal contents.",
                expected=("Scenario blocks render for SC-PP-008, SC-LN-002, "
                          "SC-PP-UBO-023 (and any others with highlights)."),
            ),
            Step(
                instructions="Inspect a scenario block's table.",
                expected=("Columns: Decision pill, Driver (ontology type), "
                          "Cases, Share of decision, Share of all, Sample "
                          "fact ids, Promote button."),
            ),
            Step(
                instructions="Scan the row ordering within a block.",
                expected="Patterns are sorted by case count desc.",
            ),
            Step(
                instructions=("Locate the SanctionsProximity row on "
                              "SC-PP-008."),
                expected=("Share of reject decisions is high (100% on the "
                          "backfilled data)."),
            ),
        ],
    ),
    TestCase(
        tc_id="TC-AD-04",
        title="Promote pattern (Phase 3b)",
        pre="Insights modal open from TC-AD-03.",
        steps=[
            Step(
                instructions=("Click Promote -> on the SC-PP-008 / reject / "
                              "SanctionsProximity row."),
                expected=("Confirmation modal appears with \"Promote this "
                          "driver to a scenario rule?\"."),
            ),
            Step(
                instructions="Read the confirmation body.",
                expected=("Text references the case count + share %. "
                          "Mentions the new version that will be created."),
            ),
            Step(
                instructions="Click Promote pattern.",
                expected=("Alert: \"Promoted. SC-PP-008 is now at vN+1. "
                          "Future cases that bind SanctionsProximity facts "
                          "will show an advisory chip...\"."),
            ),
            Step(
                instructions="Watch the Insights modal refresh.",
                expected=("Pattern still listed; now persisted on the "
                          "scenario YAML."),
            ),
            Step(
                instructions=("Open SC-PP-008 in the Edit modal -> Version "
                              "history."),
                expected=("New version visible. The YAML at that version "
                          "contains auto_promoted_patterns with the "
                          "SanctionsProximity entry."),
            ),
        ],
    ),
    TestCase(
        tc_id="TC-AD-05",
        title="Guardrail blocks low-confidence promotion",
        pre="Insights modal open.",
        steps=[
            Step(
                instructions=("Identify a pattern with case_count = 1 in the "
                              "table."),
                expected="The Promote button is still visible (UI doesn't pre-filter).",
            ),
            Step(
                instructions="Click Promote -> then Promote pattern.",
                expected=("Inline error: \"guardrail: case_count=1 < "
                          "min_case_count=2\" (the default)."),
            ),
            Step(
                instructions="Re-open the scenario in the Edit modal.",
                expected=("No new version. auto_promoted_patterns NOT changed "
                          "for this trigger."),
            ),
        ],
    ),
    TestCase(
        tc_id="TC-AD-06",
        title="Demote pattern",
        pre="TC-AD-04 promoted a pattern. You have the pattern_id from the response.",
        steps=[
            Step(
                instructions=("Curl POST /api/insights/patterns/demote with "
                              "{scenario_id, pattern_id} as admin."),
                expected=("200 OK. Response shows version bumped again "
                          "(now vN+2)."),
            ),
            Step(
                instructions="Open SC-PP-008 in the edit modal.",
                expected=("auto_promoted_patterns no longer contains the "
                          "entry."),
            ),
            Step(
                instructions="Start a brand-new SC-PP-008 case.",
                expected=("Advisory chip no longer appears in the envelope."),
            ),
            Step(
                instructions="Open the older SC-PP-008 case pinned to vN+1.",
                expected=("Older case STILL shows the chip (pinned to vN+1 "
                          "snapshot) -- by design."),
            ),
        ],
    ),
    TestCase(
        tc_id="TC-AD-07",
        title="Non-admin cannot access Insights or Promote",
        pre="Have one admin token AND one non-admin token.",
        steps=[
            Step(
                instructions="Log out, log back in as a non-admin operator.",
                expected="StatusBar does NOT show the Insights button.",
            ),
            Step(
                instructions=("Curl GET /api/insights/patterns with the "
                              "operator's bearer token."),
                expected="HTTP 403.",
            ),
            Step(
                instructions=("Curl POST /api/insights/patterns/promote "
                              "with the operator's bearer token."),
                expected="HTTP 403.",
            ),
        ],
    ),
]


# -----------------------------------------------------------------------------
# D — Auditor workflows
# -----------------------------------------------------------------------------
D_AUDITOR: list[TestCase] = [
    TestCase(
        tc_id="TC-AU-01",
        title="Trace a case to its source spec",
        pre=("A closed case exists from at least 30 days ago. The scenario "
             "it ran against has been edited since."),
        steps=[
            Step(
                instructions="Open the historical case from the case list.",
                expected="Case detail loads.",
            ),
            Step(
                instructions="Open Case spec modal.",
                expected=("Header pill reads \"Ran against vN\" where N is "
                          "earlier than the current live version."),
            ),
            Step(
                instructions="Read the Scenario tab content.",
                expected=("Content reflects vN spec -- not live."),
            ),
            Step(
                instructions="Switch to the Ontology tab.",
                expected=("Class definitions + mappings shown match the "
                          "vN snapshot."),
            ),
        ],
        pass_criteria=("Auditor can answer \"what spec was the agent running "
                       "when this decision was made?\" without git archaeology."),
    ),
    TestCase(
        tc_id="TC-AU-02",
        title="Audit-trail immutability",
        pre="A closed case from any past session.",
        steps=[
            Step(
                instructions="Open the case. Locate the Audit trail panel.",
                expected="Append-only log of stage events renders.",
            ),
            Step(
                instructions=("Scan the sequence numbers across the events."),
                expected="Sequence numbers are contiguous (no gaps).",
            ),
            Step(
                instructions=("Wait several minutes, re-open the same case "
                              "(F5 the page)."),
                expected=("Same events in the same order; nothing rewritten "
                          "or removed."),
            ),
        ],
    ),
    TestCase(
        tc_id="TC-AU-03",
        title="Verify load-bearing facts persist",
        pre=("A closed case from the 26 backfilled set or a Phase 2 capture "
             "is selected."),
        steps=[
            Step(
                instructions="Open Case spec on the case.",
                expected=("\"Reviewer marked these facts as load-bearing (N)\" "
                          "section is visible. N matches the demo seed count."),
            ),
            Step(
                instructions=("Curl GET /api/cases/{id}/highlights with an "
                              "admin token."),
                expected=("Response returns the same refs as the UI in "
                          "the same order."),
            ),
            Step(
                instructions=("Inspect each ref in the response: source, "
                              "ontology_type, id, position."),
                expected=("All fields populated. Position numbers are 0..N-1."),
            ),
        ],
    ),
    TestCase(
        tc_id="TC-AU-04",
        title="Verify promoted-pattern provenance",
        pre=("A scenario has a promoted pattern (e.g. SC-PP-008 after "
             "TC-AD-04)."),
        steps=[
            Step(
                instructions=("Open the scenario YAML at the post-promotion "
                              "version (Version history -> latest)."),
                expected=("auto_promoted_patterns entry is visible with "
                          "metadata fields."),
            ),
            Step(
                instructions=("Inspect metadata.promoted_by, promoted_at, "
                              "source_case_count, source_share_of_kind_pct, "
                              "source_total_decided."),
                expected=("All five fields populated. Auditor can answer "
                          "\"who promoted this rule, when, based on what?\"."),
            ),
        ],
    ),
]


# -----------------------------------------------------------------------------
# E — End-to-end demo flows
# -----------------------------------------------------------------------------
E_DEMO: list[TestCase] = [
    TestCase(
        tc_id="TC-DEMO-01",
        title="The compounding loop, one cycle",
        pre=("Three operator-style and admin-style accounts available. Demo "
             "audience present. Tell the platform's story end-to-end."),
        steps=[
            Step(
                instructions=("Operator: submit \"onboard Hemlock Precision "
                              "Castings\"."),
                expected=("SC-PP-008 fires at v1. ~13 facts surface. "
                          "Routed to reviewer."),
            ),
            Step(
                instructions=("Reviewer: reject, pick SanctionsProximity "
                              "as load-bearing."),
                expected=("Case completes as reject. Phase 2 capture "
                          "recorded."),
            ),
            Step(
                instructions=("Repeat steps 1-2 with two more prompts so "
                              "three rejects all citing SanctionsProximity "
                              "exist."),
                expected=("Three closed reject cases on SC-PP-008. "
                          "case_highlighted_facts grew by 3."),
            ),
            Step(
                instructions=("Admin: open Insights."),
                expected=("Row visible: \"SC-PP-008 / reject / "
                          "SanctionsProximity / 3 cases / 100%\"."),
            ),
            Step(
                instructions="Admin: click Promote -> then Promote pattern.",
                expected=("Alert confirms SC-PP-008 bumped to v2 with "
                          "auto_promoted_patterns entry."),
            ),
            Step(
                instructions=("Operator: submit a fourth supplier-onboarding "
                              "prompt."),
                expected=("Case enters Awaiting Review."),
            ),
            Step(
                instructions=("Once binding completes, view the envelope."),
                expected=("Advisory banner appears: \"Compounding loop * 1 "
                          "promoted pattern matched * REJECT * "
                          "SanctionsProximity * Reviewers cited "
                          "SanctionsProximity in 3 of 3 reject decisions...\". "
                          "Numbers match steps 1-3."),
            ),
            Step(
                instructions=("Reviewer: see the chip, decide normally "
                              "(approve or reject)."),
                expected=("Reviewer is not coerced. Audit trail records "
                          "who promoted, when, from how many cases."),
            ),
        ],
        pass_criteria=("Every step renders correctly. The chip in step 7 "
                       "cites the actual numbers from steps 1-3. The platform "
                       "demonstrably learned from reviewer behaviour."),
    ),
    TestCase(
        tc_id="TC-DEMO-02",
        title="Mode-switch case (SC-LN-002)",
        pre="Operator logged in. Logistics demo data seeded.",
        steps=[
            Step(
                instructions=("Submit: \"switch shipment S-700412 from ocean "
                              "to air to recover the schedule.\""),
                expected="SC-LN-002 picks up.",
            ),
            Step(
                instructions=("Inspect the envelope after binding."),
                expected=("~13 facts including alternate carrier options "
                          "(FreightRate cards)."),
            ),
            Step(
                instructions=("Reviewer (planner): approve. Confirm execute."),
                expected=("Executor runs against logistics.sqlite. Closing "
                          "message shows before/after diff."),
            ),
            Step(
                instructions="Re-open via Case spec.",
                expected=("Pill: \"Ran against v1\". Highlights section "
                          "renders if applicable."),
            ),
        ],
    ),
    TestCase(
        tc_id="TC-DEMO-03",
        title="Autonomous case (SC-LN-DEMURRAGE-AUTO-025)",
        pre="Operator logged in.",
        steps=[
            Step(
                instructions=("Submit: \"auto-dispute demurrage charge "
                              "around $300 for shipment X.\""),
                expected=("Agent picks the demurrage auto-dispute scenario."),
            ),
            Step(
                instructions=("Watch the case progress without HITL."),
                expected=("Guardrail policy fires -> auto_approve. Case "
                          "moves straight to Executed."),
            ),
            Step(
                instructions=("Inspect the closing message + execution result."),
                expected=("Pre / post state visible in the closing message. "
                          "No reviewer interaction required."),
            ),
            Step(
                instructions="Open the Audit trail panel.",
                expected=("\"auto-approved\" event present with the "
                          "guardrail id."),
            ),
        ],
    ),
]


SECTIONS: list[tuple[str, list[TestCase]]] = [
    ("A_Operator", A_OPERATOR),
    ("B_Reviewer", B_REVIEWER),
    ("C_Admin", C_ADMIN),
    ("D_Auditor", D_AUDITOR),
    ("E_Demo", E_DEMO),
]


# -----------------------------------------------------------------------------
# Workbook generation
# -----------------------------------------------------------------------------
HEADER = ["TC-ID", "Test", "Instructions", "Expected outcome", "Pass/Fail", "Notes"]

HEADER_FILL = PatternFill(start_color="FF4F46E5", end_color="FF4F46E5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
SETUP_FILL = PatternFill(start_color="FFEEF2FF", end_color="FFEEF2FF", fill_type="solid")
SETUP_FONT = Font(italic=True, color="FF4338CA")
PASS_FILL = PatternFill(start_color="FFFEF3C7", end_color="FFFEF3C7", fill_type="solid")
PASS_FONT = Font(bold=True, color="FF92400E")

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _write_header(ws: Worksheet) -> None:
    for col_idx, label in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    widths = {"A": 12, "B": 36, "C": 50, "D": 52, "E": 12, "F": 28}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_tc(ws: Worksheet, start_row: int, tc: TestCase) -> int:
    """Write one test case starting at `start_row`. Returns the next free row."""
    row = start_row

    # Setup row
    setup_test = f"{tc.title} - Setup"
    cells = [
        (tc.tc_id, None),
        (setup_test, SETUP_FONT),
        (tc.pre, SETUP_FONT),
        ("-", SETUP_FONT),
        ("-", SETUP_FONT),
        ("", None),
    ]
    for col_idx, (val, font) in enumerate(cells, start=1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.alignment = WRAP
        c.fill = SETUP_FILL
        if font:
            c.font = font
    row += 1

    # Step rows
    for i, step in enumerate(tc.steps, start=1):
        ws.cell(row=row, column=1, value=tc.tc_id).alignment = WRAP
        ws.cell(row=row, column=2, value=f"{tc.title} - Step {i}").alignment = WRAP
        ws.cell(row=row, column=3, value=step.instructions).alignment = WRAP
        ws.cell(row=row, column=4, value=step.expected).alignment = WRAP
        ws.cell(row=row, column=5, value="").alignment = WRAP_CENTER
        ws.cell(row=row, column=6, value="").alignment = WRAP
        row += 1

    # Pass criteria rollup (optional)
    if tc.pass_criteria:
        cells = [
            (tc.tc_id, None),
            (f"{tc.title} - Pass criteria", PASS_FONT),
            ("(rollup across the steps above)", None),
            (tc.pass_criteria, PASS_FONT),
            ("", None),
            ("", None),
        ]
        for col_idx, (val, font) in enumerate(cells, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.alignment = WRAP
            c.fill = PASS_FILL
            if font:
                c.font = font
        row += 1

    return row


def build() -> Path:
    wb = Workbook()
    # Remove the default sheet that openpyxl creates.
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    for sheet_name, test_cases in SECTIONS:
        ws = wb.create_sheet(title=sheet_name)
        _write_header(ws)
        row = 2
        for tc in test_cases:
            row = _write_tc(ws, row, tc)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    return OUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"wrote {path}")
    # Summary so the operator running the script can spot-check.
    total_rows = 0
    for name, cases in SECTIONS:
        rows = sum(1 + len(tc.steps) + (1 if tc.pass_criteria else 0) for tc in cases)
        print(f"  {name}: {len(cases)} test case(s), {rows} data row(s)")
        total_rows += rows
    print(f"  total data rows across all sheets: {total_rows}")
